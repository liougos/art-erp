import os
import re
import json
import base64
import logging
import tempfile
from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, send_from_directory, jsonify
from flask_login import login_required, current_user
from datetime import datetime, date
from werkzeug.utils import secure_filename
from models import db, Invoice, Project, InvoicePayment, BankAccount

logger = logging.getLogger(__name__)

invoices_bp = Blueprint('invoices', __name__)

ALLOWED_IMG = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'heic'}
ALLOWED_ALL  = ALLOWED_IMG | {'pdf'}

CATEGORIES = ['Υλικά Συντήρησης', 'Εργασία', 'Εξοπλισμός', 'Υπεργολαβία',
              'Ενοίκιο Σκαλωσιάς', 'Μεταφορικά', 'Ξενοδοχεία/Αποζημιώσεις',
              'Γενικά Έξοδα', 'ΔΕΗ/ΟΤΕ', 'Καύσιμα', 'Αναλώσιμα', 'Άλλο']


def _sanitize(name: str) -> str:
    """Transliterate Greek and keep only safe filename chars."""
    GR = str.maketrans(
        'ΑΒΓΔΕΖΗΘΙΚΛΜΝΞΟΠΡΣΤΥΦΧΨΩαβγδεζηθικλμνξοπρστυφχψω',
        'AVGDEZHTHIKLMNXOPRSTYFXPOavgdezithiklmnxoprstyfxpo'
    )
    name = name.translate(GR)
    name = re.sub(r'[^\w\-]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name[:60] or 'τιμολογιο'


def _image_to_pdf(img_path: str, pdf_path: str):
    """Convert an image file to a PDF using Pillow."""
    from PIL import Image
    img = Image.open(img_path)
    if img.mode in ('RGBA', 'LA', 'P', 'CMYK'):
        img = img.convert('RGB')
    img.save(pdf_path, 'PDF', resolution=150, save_all=False)


def _save_invoice_file(file, store_name: str, inv_date: date, upload_dir: str):
    """Save uploaded file, converting image to PDF. Returns (stored_path, original_name)."""
    os.makedirs(upload_dir, exist_ok=True)
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    date_str = inv_date.strftime('%d_%m_%Y')
    base = f"{_sanitize(store_name)}_{date_str}"
    is_image = ext in ALLOWED_IMG

    if is_image:
        # Save original image first, then convert
        tmp_name = f"_tmp_{base}.{ext}"
        tmp_path = os.path.join(upload_dir, tmp_name)
        file.save(tmp_path)
        pdf_name = f"{base}.pdf"
        pdf_path = os.path.join(upload_dir, pdf_name)
        _image_to_pdf(tmp_path, pdf_path)
        os.remove(tmp_path)
        return pdf_name, file.filename
    else:
        # PDF or other — keep as-is
        pdf_name = f"{base}.pdf"
        file.save(os.path.join(upload_dir, pdf_name))
        return pdf_name, file.filename


@invoices_bp.route('/')
@login_required
def index():
    inv_type = request.args.get('type', '')
    status = request.args.get('status', '')
    project_id = request.args.get('project_id', '')
    q = request.args.get('q', '')

    query = Invoice.query
    if inv_type: query = query.filter_by(invoice_type=inv_type)
    if status: query = query.filter_by(payment_status=status)
    if project_id: query = query.filter_by(project_id=int(project_id))
    if q: query = query.filter(
        Invoice.invoice_number.ilike(f'%{q}%') | Invoice.issuer.ilike(f'%{q}%')
    )

    invoices = query.order_by(Invoice.invoice_date.desc()).all()
    projects = Project.query.order_by(Project.code).all()

    total_income  = sum(float(i.total_amount or 0) for i in invoices if i.invoice_type == 'income')
    total_expense = sum(float(i.total_amount or 0) for i in invoices if i.invoice_type == 'expense')
    overdue = [i for i in invoices if i.is_overdue]

    return render_template('invoices/index.html', invoices=invoices, projects=projects,
                           total_income=total_income, total_expense=total_expense,
                           overdue=overdue, inv_type=inv_type, status=status,
                           project_id=project_id, q=q, categories=CATEGORIES)


@invoices_bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    projects = Project.query.filter(Project.status.in_(['active', 'planning'])).order_by(Project.code).all()
    if request.method == 'POST':
        store_name = request.form.get('issuer', 'τιμολόγιο').strip()
        inv_date_str = request.form.get('invoice_date', str(date.today()))
        try:
            inv_date = datetime.strptime(inv_date_str, '%Y-%m-%d').date()
        except ValueError:
            inv_date = date.today()

        image_path = None
        file = request.files.get('invoice_image')
        if file and file.filename:
            ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
            if ext in ALLOWED_ALL:
                upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'invoices')
                fname, orig_name = _save_invoice_file(file, store_name, inv_date, upload_dir)
                image_path = f'invoices/{fname}'

        net = float(request.form.get('amount_net', 0) or 0)
        vat_rate = float(request.form.get('vat_rate', 24) or 24)
        vat = round(net * vat_rate / 100, 2)
        total = round(net + vat, 2)

        payment_method = request.form.get('payment_method', '').strip()

        inv = Invoice(
            invoice_number=request.form.get('invoice_number', '').strip(),
            invoice_type=request.form.get('invoice_type', 'expense'),
            issuer=store_name,
            issuer_afm=request.form.get('issuer_afm', '').strip(),
            recipient=request.form.get('recipient', 'ART RESTORATION AE').strip(),
            amount_net=net,
            vat_rate=vat_rate,
            vat_amount=vat,
            total_amount=total,
            category=request.form.get('category', '').strip(),
            project_id=request.form.get('project_id') or None,
            payment_status=request.form.get('payment_status', 'pending'),
            payment_method=payment_method,
            image_path=image_path,
            source='photo' if image_path else 'manual',
            notes=request.form.get('notes', '').strip(),
            uploaded_by_id=current_user.id,
            invoice_date=inv_date,
        )
        dd = request.form.get('due_date')
        if dd: inv.due_date = datetime.strptime(dd, '%Y-%m-%d').date()
        pd = request.form.get('payment_date')
        if pd: inv.payment_date = datetime.strptime(pd, '%Y-%m-%d').date()

        db.session.add(inv)
        db.session.commit()

        # Upload to Google Drive
        if image_path:
            try:
                from services.google_drive import upload_file as drive_upload
                local_path = os.path.join(current_app.config['UPLOAD_FOLDER'], image_path)
                year_folder = str(inv_date.year)
                drive_upload(local_path, 'invoices', subfolder=year_folder, filename=os.path.basename(image_path))
            except Exception as e:
                logger.warning(f'Drive upload skipped: {e}')

        # Notify admin if payment is on credit
        if payment_method.lower() in ('πίστωση', 'credit', 'πιστωση', 'επί πιστώσει'):
            from models import Notification, User as _User
            admins = _User.query.filter(_User.role.in_(['admin', 'manager']), _User.is_active == True).all()
            for admin in admins:
                n = Notification(
                    user_id=admin.id,
                    title=f'Πληρωμή επί πιστώσει — {store_name}',
                    message=f'Τιμολόγιο {total:.2f}€ από {store_name} — απαιτείται εξόφληση.',
                    link='/invoices/',
                    icon='bi-credit-card',
                )
                db.session.add(n)
            db.session.commit()
            flash(f'Τιμολόγιο καταχωρήθηκε. Ειδοποίηση για πληρωμή πίστωσης στάλθηκε.', 'warning')
        else:
            flash(f'Τιμολόγιο {inv.invoice_number or "#" + str(inv.id)} καταχωρήθηκε.', 'success')

        return redirect(url_for('invoices.index'))
    return render_template('invoices/upload.html', projects=projects, categories=CATEGORIES, today=date.today())


@invoices_bp.route('/view/<int:id>')
@login_required
def view_file(id):
    inv = Invoice.query.get_or_404(id)
    if inv.image_path:
        directory = os.path.join(current_app.config['UPLOAD_FOLDER'],
                                 os.path.dirname(inv.image_path.replace('invoices/', '')))
        fname = os.path.basename(inv.image_path)
        return send_from_directory(
            os.path.join(current_app.config['UPLOAD_FOLDER'], 'invoices'),
            fname, as_attachment=False
        )
    flash('Δεν υπάρχει αρχείο.', 'warning')
    return redirect(url_for('invoices.index'))


@invoices_bp.route('/<int:id>/pay', methods=['POST'])
@login_required
def mark_paid(id):
    inv = Invoice.query.get_or_404(id)
    inv.payment_status = 'paid'
    inv.payment_date = date.today()
    inv.payment_method = request.form.get('payment_method', 'Τραπεζική μεταφορά')
    db.session.commit()
    flash('Το τιμολόγιο σημειώθηκε ως πληρωμένο.', 'success')
    return redirect(url_for('invoices.index'))


@invoices_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    inv = Invoice.query.get_or_404(id)
    db.session.delete(inv)
    db.session.commit()
    flash('Τιμολόγιο διαγράφηκε.', 'success')
    return redirect(url_for('invoices.index'))


@invoices_bp.route('/<int:id>')
@login_required
def detail(id):
    inv = Invoice.query.get_or_404(id)
    bank_accounts = BankAccount.query.filter_by(is_active=True).order_by(BankAccount.name).all()
    return render_template('invoices/detail.html', inv=inv, bank_accounts=bank_accounts, today=date.today())


@invoices_bp.route('/<int:id>/payment', methods=['POST'])
@login_required
def add_payment(id):
    inv = Invoice.query.get_or_404(id)
    try:
        amount = float(request.form['amount'])
        if amount <= 0:
            flash('Το ποσό πρέπει να είναι θετικό.', 'danger')
            return redirect(url_for('invoices.detail', id=id))
        if amount > inv.remaining_balance + 0.01:
            flash(f'Το ποσό ({amount:.2f}€) υπερβαίνει το υπόλοιπο ({inv.remaining_balance:.2f}€).', 'warning')
            return redirect(url_for('invoices.detail', id=id))

        pmt = InvoicePayment(
            invoice_id=id,
            amount=amount,
            payment_method=request.form.get('payment_method', '').strip(),
            bank_account_id=request.form.get('bank_account_id') or None,
            reference=request.form.get('reference', '').strip(),
            notes=request.form.get('notes', '').strip(),
            created_by_id=current_user.id,
        )
        pd = request.form.get('payment_date')
        pmt.payment_date = datetime.strptime(pd, '%Y-%m-%d').date() if pd else date.today()
        db.session.add(pmt)

        # Update invoice payment_status
        new_paid = inv.total_paid + amount
        total = float(inv.total_amount or 0)
        if new_paid >= total - 0.01:
            inv.payment_status = 'paid'
            inv.payment_date = pmt.payment_date
        else:
            inv.payment_status = 'partial'
        db.session.commit()
        flash(f'Πληρωμή {amount:.2f}€ καταχωρήθηκε. Υπόλοιπο: {max(0, total - new_paid):.2f}€', 'success')
    except Exception as e:
        flash(f'Σφάλμα: {e}', 'danger')
    return redirect(url_for('invoices.detail', id=id))


@invoices_bp.route('/<int:id>/payment/<int:pid>/delete', methods=['POST'])
@login_required
def delete_payment(id, pid):
    if current_user.role not in ('admin', 'manager'):
        flash('Δεν έχετε δικαίωμα.', 'danger')
        return redirect(url_for('invoices.detail', id=id))
    pmt = InvoicePayment.query.filter_by(id=pid, invoice_id=id).first_or_404()
    db.session.delete(pmt)
    inv = Invoice.query.get(id)
    # Recalculate status
    new_paid = sum(float(p.amount) for p in inv.payments if p.id != pid)
    total = float(inv.total_amount or 0)
    inv.payment_status = 'paid' if new_paid >= total - 0.01 else ('partial' if new_paid > 0 else 'pending')
    db.session.commit()
    flash('Πληρωμή διαγράφηκε.', 'success')
    return redirect(url_for('invoices.detail', id=id))


@invoices_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    inv = Invoice.query.get_or_404(id)
    projects = Project.query.filter(Project.status.in_(['active', 'planning'])).order_by(Project.code).all()
    if request.method == 'POST':
        inv.invoice_number = request.form.get('invoice_number', '').strip()
        inv.invoice_type   = request.form.get('invoice_type', inv.invoice_type)
        inv.issuer         = request.form.get('issuer', '').strip()
        inv.issuer_afm     = request.form.get('issuer_afm', '').strip()
        inv.recipient      = request.form.get('recipient', '').strip()
        inv.category       = request.form.get('category', '').strip()
        inv.project_id     = request.form.get('project_id') or None
        inv.notes          = request.form.get('notes', '').strip()
        net = float(request.form.get('amount_net', inv.amount_net) or 0)
        vat_rate = float(request.form.get('vat_rate', inv.vat_rate) or 24)
        inv.amount_net = net
        inv.vat_rate = vat_rate
        inv.vat_amount = round(net * vat_rate / 100, 2)
        inv.total_amount = round(net + inv.vat_amount, 2)
        idate = request.form.get('invoice_date')
        if idate: inv.invoice_date = datetime.strptime(idate, '%Y-%m-%d').date()
        ddate = request.form.get('due_date')
        inv.due_date = datetime.strptime(ddate, '%Y-%m-%d').date() if ddate else None
        db.session.commit()
        flash('Τιμολόγιο ενημερώθηκε.', 'success')
        return redirect(url_for('invoices.detail', id=id))
    return render_template('invoices/edit.html', inv=inv, projects=projects, categories=CATEGORIES, today=date.today())


# ── AI EXTRACTION ───────────────────────────────────────────────────────────

_AI_PROMPT = """Ανάλυσε αυτό το ελληνικό τιμολόγιο και επέστρεψε ΜΟΝΟ έγκυρο JSON (χωρίς markdown, χωρίς εξηγήσεις):
{
  "invoice_number": "αριθμός τιμολογίου ή κενή string",
  "invoice_date": "YYYY-MM-DD ή κενή string",
  "issuer": "επωνυμία εκδότη",
  "issuer_afm": "ΑΦΜ εκδότη ή κενή string",
  "amount_net": αριθμός_χωρίς_σύμβολο,
  "vat_rate": αριθμός (0, 6, 13, ή 24),
  "vat_amount": αριθμός,
  "total_amount": αριθμός,
  "invoice_type": "expense"
}
Χρησιμοποίησε τελεία ως δεκαδικό (όχι κόμμα). Αν δεν βρεις κάτι βάλε κενή string ή 0."""


_EMPTY_EXTRACTION = {
    'invoice_number': '', 'invoice_date': '', 'issuer': '', 'issuer_afm': '',
    'amount_net': 0, 'vat_rate': 24, 'vat_amount': 0, 'total_amount': 0,
    'invoice_type': 'expense', '_partial': True,
}


@invoices_bp.route('/extract', methods=['POST'])
@login_required
def extract():
    """
    AI extraction — NEVER returns a non-200 or hard error.
    On any failure returns _EMPTY_EXTRACTION so the user can fill manually.
    """
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify(_EMPTY_EXTRACTION)

    # Check API key up front — fail gracefully, not with 500
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        logger.warning('ANTHROPIC_API_KEY not set — returning empty extraction')
        return jsonify(dict(_EMPTY_EXTRACTION, _reason='no_api_key'))

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)

        if ext == 'pdf':
            data = file.read()
            # First try text extraction (fast, free)
            text = ''
            try:
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                    tmp.write(data)
                    tmp_path = tmp.name
                from pypdf import PdfReader
                reader = PdfReader(tmp_path)
                text = '\n'.join(p.extract_text() or '' for p in reader.pages[:10])
                os.unlink(tmp_path)
            except Exception as pdf_err:
                logger.debug('PDF text extraction failed: %s', pdf_err)

            if text.strip():
                msg = client.messages.create(
                    model='claude-opus-4-5', max_tokens=1024,
                    messages=[{'role': 'user', 'content': f'{_AI_PROMPT}\n\nΚείμενο τιμολογίου:\n{text[:8000]}'}]
                )
            else:
                # PDF has no extractable text (scanned) → use vision via base64
                b64 = base64.standard_b64encode(data).decode()
                msg = client.messages.create(
                    model='claude-opus-4-5', max_tokens=1024,
                    messages=[{'role': 'user', 'content': [
                        {'type': 'document', 'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': b64}},
                        {'type': 'text', 'text': _AI_PROMPT}
                    ]}]
                )
        else:
            data = file.read()
            b64 = base64.standard_b64encode(data).decode()
            mt = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png',
                  'gif': 'image/gif', 'webp': 'image/webp', 'heic': 'image/heic'}.get(ext, 'image/jpeg')
            msg = client.messages.create(
                model='claude-opus-4-5', max_tokens=1024,
                messages=[{'role': 'user', 'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': mt, 'data': b64}},
                    {'type': 'text', 'text': _AI_PROMPT}
                ]}]
            )

        raw = msg.content[0].text.strip()
        # Try to parse JSON — handle markdown code blocks too
        raw_clean = re.sub(r'^```(?:json)?\s*|\s*```$', '', raw, flags=re.MULTILINE).strip()
        match = re.search(r'\{.*\}', raw_clean, re.DOTALL)
        if match:
            result = json.loads(match.group())
            # Ensure all expected keys exist
            for k, v in _EMPTY_EXTRACTION.items():
                result.setdefault(k, v)
            return jsonify(result)

        # AI responded but no JSON found — return empty
        logger.warning('AI extraction: no JSON in response: %s', raw[:200])
        return jsonify(_EMPTY_EXTRACTION)

    except Exception as e:
        logger.error('Invoice extraction error: %s', e)
        # Always 200 — frontend handles partial gracefully
        return jsonify(dict(_EMPTY_EXTRACTION, _reason=str(e)[:120]))


# ── ACCOUNTANT LIST IMPORT ──────────────────────────────────────────────────

def _parse_import_row(r: dict):
    def get(*keys):
        for k in keys:
            v = r.get(k, '') or r.get(k.upper(), '') or r.get(k.capitalize(), '')
            if v and str(v).strip() not in ('', 'None', 'nan'): return str(v).strip()
        return ''

    issuer = get('εκδότης', 'επωνυμια', 'επωνυμία', 'issuer', 'supplier', 'vendor',
                 'προμηθευτης', 'προμηθευτής', 'εταιρεια', 'εταιρεία')
    if not issuer:
        return None

    date_str = get('ημερομηνια', 'ημερομηνία', 'date', 'ημ/νια', 'ημ. τιμολογίου',
                   'invoice_date', 'ημ.τιμ.', 'ημ.τιμολογιου')
    inv_date = ''
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y', '%m/%d/%Y'):
        try:
            inv_date = datetime.strptime(date_str.strip(), fmt).strftime('%Y-%m-%d')
            break
        except Exception:
            pass

    def to_float(s):
        if not s: return 0.0
        s = str(s).replace(',', '.').replace('€', '').replace(' ', '').strip()
        try: return float(s)
        except: return 0.0

    net = to_float(get('καθαρο', 'καθαρό', 'καθαρη αξια', 'net', 'amount_net',
                        'αξια', 'αξία', 'καθαρο ποσο', 'αξια χωρις φπα'))
    total = to_float(get('συνολο', 'σύνολο', 'total', 'total_amount', 'συνολικη αξια',
                          'πληρωτεο', 'γενικο συνολο', 'γενικό σύνολο'))
    vat_rate = to_float(get('φπα%', 'φπα %', 'vat_rate', 'vat%', 'συντελεστης φπα', 'φπα')) or 24.0
    if vat_rate <= 1:
        vat_rate = round(vat_rate * 100, 0)

    if not net and total:
        net = round(total / (1 + vat_rate / 100), 2)
    vat_amount = round(net * vat_rate / 100, 2)
    if not total:
        total = round(net + vat_amount, 2)

    return {
        'invoice_number': get('αριθμος', 'αριθμός', 'αρ. τιμολογίου', 'invoice_number',
                               'αρ.τιμ.', 'number', 'αρ'),
        'invoice_date': inv_date,
        'issuer': issuer,
        'issuer_afm': get('αφμ', 'afm', 'vat_number', 'tax_id', 'α.φ.μ.'),
        'amount_net': net,
        'vat_rate': vat_rate,
        'vat_amount': vat_amount,
        'total_amount': total,
        'category': get('κατηγορια', 'κατηγορία', 'category'),
    }


@invoices_bp.route('/import-list', methods=['GET', 'POST'])
@login_required
def import_list():
    if request.method == 'GET':
        projects = Project.query.filter(Project.status.in_(['active', 'planning'])).order_by(Project.code).all()
        return render_template('invoices/import_list.html', categories=CATEGORIES, projects=projects)

    file = request.files.get('list_file')
    if not file or not file.filename:
        flash('Δεν επιλέχτηκε αρχείο.', 'danger')
        return redirect(request.url)

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    rows = []
    try:
        if ext in ('xlsx', 'xls'):
            import openpyxl
            data = file.read()
            with tempfile.NamedTemporaryFile(suffix=f'.{ext}', delete=False) as tmp:
                tmp.write(data)
                tmp_path = tmp.name
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            os.unlink(tmp_path)
            headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None):
                    continue
                rows.append(dict(zip(headers, [str(v if v is not None else '').strip() for v in row])))
        elif ext == 'csv':
            import csv, io
            text = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(text))
            rows = [{k.strip().lower(): (v or '').strip() for k, v in r.items()} for r in reader]
        else:
            flash('Αποδεκτά αρχεία: xlsx, xls, csv', 'danger')
            return redirect(request.url)
    except Exception as e:
        flash(f'Σφάλμα ανάγνωσης αρχείου: {e}', 'danger')
        return redirect(request.url)

    parsed = [p for p in (_parse_import_row(r) for r in rows) if p]
    if not parsed:
        flash('Δεν βρέθηκαν έγκυρες εγγραφές. Βεβαιωθείτε ότι το αρχείο έχει στήλη "Εκδότης" ή "Επωνυμία".', 'warning')
        return redirect(request.url)

    # ── Match each row against existing invoices in the DB ──────────────────
    existing = Invoice.query.filter_by(invoice_type='expense').all()

    def _norm(s):
        return str(s or '').strip().lower().replace(' ', '')

    for row in parsed:
        row['status'] = 'missing'
        row['match_id'] = None
        inv_no = _norm(row.get('invoice_number', ''))
        issuer  = _norm(row.get('issuer', ''))[:12]
        amount  = float(row.get('total_amount', 0) or 0)
        r_date  = row.get('invoice_date', '')

        for inv in existing:
            e_no     = _norm(inv.invoice_number or '')
            e_issuer = _norm(inv.issuer or '')[:12]
            e_amount = float(inv.total_amount or 0)
            e_date   = str(inv.invoice_date) if inv.invoice_date else ''

            # Match by invoice number + issuer prefix
            if inv_no and e_no and inv_no == e_no and issuer and issuer == e_issuer:
                row['status'] = 'found'
                row['match_id'] = inv.id
                break
            # Match by issuer + date + amount (±1€)
            if issuer and issuer == e_issuer and r_date and r_date == e_date and abs(amount - e_amount) < 1.0:
                row['status'] = 'found'
                row['match_id'] = inv.id
                break

    projects = Project.query.filter(Project.status.in_(['active', 'planning'])).order_by(Project.code).all()
    return render_template('invoices/import_preview.html',
                           parsed=parsed, projects=projects, categories=CATEGORIES,
                           rows_json=json.dumps(parsed, ensure_ascii=False))


@invoices_bp.route('/import-confirm', methods=['POST'])
@login_required
def import_confirm():
    rows_json = request.form.get('rows_json', '[]')
    try:
        rows = json.loads(rows_json)
    except Exception:
        flash('Σφάλμα δεδομένων εισαγωγής.', 'danger')
        return redirect(url_for('invoices.import_list'))

    project_id = request.form.get('project_id') or None
    selected = set(request.form.getlist('selected_rows'))
    count = 0
    for i, row in enumerate(rows):
        if str(i) not in selected:
            continue
        try:
            inv = Invoice(
                invoice_number=row.get('invoice_number', ''),
                invoice_type='expense',
                issuer=row.get('issuer', ''),
                issuer_afm=row.get('issuer_afm', ''),
                recipient='ART RESTORATION AE',
                amount_net=row.get('amount_net', 0),
                vat_rate=row.get('vat_rate', 24),
                vat_amount=row.get('vat_amount', 0),
                total_amount=row.get('total_amount', 0),
                category=row.get('category', ''),
                project_id=project_id,
                payment_status='pending',
                source='import',
                uploaded_by_id=current_user.id,
            )
            if row.get('invoice_date'):
                inv.invoice_date = datetime.strptime(row['invoice_date'], '%Y-%m-%d').date()
            db.session.add(inv)
            count += 1
        except Exception as e:
            logger.error(f'Import row {i} error: {e}')
    db.session.commit()
    flash(f'{count} τιμολόγια εισήχθηκαν επιτυχώς.', 'success')
    return redirect(url_for('invoices.index'))
